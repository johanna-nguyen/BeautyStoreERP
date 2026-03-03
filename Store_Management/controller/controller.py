import csv
from tkinter import messagebox, filedialog

from fpdf import FPDF
from openpyxl import Workbook


class Controller:

    def __init__(self, model, view):
        self.view = view
        self.model = model

    def hide_buttons_for_staff(self, user_role):
        """Ẩn các nút Sửa và Xóa nếu người dùng là staff"""
        pass

    def clear_treeview(self):
        """Xóa tất cả dữ liệu trên bảng Treeview"""
        pass

    def click_print(self, title, rows, headers):
        """Hàm in PDF tổng quát"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")]
            )
            if not file_path:
                return

            if not rows:
                messagebox.showinfo("Notification", "No data to print")
                return
            if not headers:
                messagebox.showerror("Error", "No column headers")
                return

            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            pdf.add_font('TimesNewRoman', '', 'fonts/TimesNewRoman.ttf', uni=True)
            pdf.set_font('TimesNewRoman', '', 18)
            pdf.cell(0, 10, title, ln=True, align="C")
            pdf.ln(10)
            pdf.set_font('TimesNewRoman', '', 8)

            col_width = 190 // len(headers)

            # Header
            for header in headers:
                pdf.cell(col_width, 8, str(header), border=1, align="C")
            pdf.ln()

            # Data
            for row in rows:
                for value in row:
                    pdf.cell(col_width, 8, str(value), border=1)
                pdf.ln()

            pdf.output(file_path)
            messagebox.showinfo("Notification", "Printed successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to print: {e}")

    def click_csv(self, title, rows, headers):
        """Xuất file CSV từ headers và rows"""
        try:
            # Mở hộp thoại lưu file
            file_path = filedialog.asksaveasfilename(
                defaultextension=".csv",
                filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            )
            if not file_path:
                return

            # Kiểm tra dữ liệu
            if not headers or not rows:
                messagebox.showinfo("Notification", "No data to export")
                return

            # Ghi dữ liệu vào file CSV
            with open(file_path, mode="w", newline="", encoding="utf-8-sig") as file:
                writer = csv.writer(file)
                writer.writerow([title]) # Ghi title
                writer.writerow(headers)  # Ghi header
                for row in rows:
                    writer.writerow(row)  # Ghi từng dòng dữ liệu

            messagebox.showinfo("Notification", "Exported successfully")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")

    def click_excel(self, title, rows, headers):
        """Xuất file Excel từ headers và rows"""
        try:
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            )
            if not file_path:
                return

            if not headers or not rows:
                messagebox.showinfo("Notification", "No data to export")
                return

            # Tạo workbook
            workbook = Workbook()
            sheet = workbook.active

            # Ghi title (dòng đầu)
            sheet.append([title])

            # Ghi header
            sheet.append(headers)

            # Ghi dữ liệu
            for row in rows:
                sheet.append(row)

            # Tự động chỉnh độ rộng cột
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                sheet.column_dimensions[column].width = max_length + 10

            workbook.save(file_path)

            messagebox.showinfo("Notification", "Exported successfully")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export: {e}")